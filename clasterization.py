import os
from pprint import pprint

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook

from helping_functions import json_work, split_list, create_excel

data = json_work("other_files/main.json", "r")


class Clasterization:
    def __init__(self, work_file, sheet_name="sheet", name_doc="file_"):
        self.index = 0
        self.workbook = None
        self.sheet = None
        self.work_file = work_file
        self.list_query = []
        self.name_doc = name_doc
        self.blacklist = []

        # self.url = url

    # def set_cluster(self, list_in, lvl):
    #     tmp_list = []  # входящие в кластер
    #     for query in list_in:
    #         query_urls = self.get_data(query)["SERP"]["url"]
    #         for query1 in list_in:
    #             query_urls1 = self.get_data(query1)["SERP"]["url"]
    #
    #             if len(set(query_urls) & set(query_urls1)) >= 7 and query1 not in tmp_list:  # если 7+ общих
    #                 tmp_list.append(query1)  # добавляем в кластер
    #
    #         tmp_list.append("/")  # отделение кластера
    #     return tmp_list

    # Получаем список запросов
    def get_dict_from_work(self):
        for item in self.work_file:
            self.list_query.append(item["maska"]["with_minsk"])

    def set_cluster(self, list_in):
        # count = 0
        # total = sum(range(0, len(list_in)))
        tmp_list = []  # входящие в кластер
        for idx, query in enumerate(list_in[:len(list_in) - 1]):
            query_urls = self.get_data(query)["SERP"]["url"]
            for query1 in list_in[idx + 1:]:
                query_urls1 = self.get_data(query1)["SERP"]["url"]
                # count += 1        # минимальное количество проверок = сумме всех чисел до значения кол-ва эл-ов
                # print(f'{count}/{total} Общих URL между {query} и {query1}: {len(set(query_urls) & set(query_urls1))}')
                if len(set(query_urls) & set(query_urls1)) >= 7 and query1 not in tmp_list:  # если 7+ общих
                    tmp_list.append(query1)  # добавляем в кластер
            tmp_list.append("/")  # отделение кластера
        return tmp_list

    def nan_delete(self, ser):
        idx = 0
        res = []
        while ser[idx] == ser[idx] and idx < ser.size - 1:
            res.append(ser[idx])
            idx += 1
        return res

    def cluster_to_excel(self, cluster, cluster_lvl, index):
        # print(cluster)
        # print(f'Cluster lvl: {cluster_lvl}')
        # print(size)
        const = 10
        sheet = self.sheet
        sheet.merge_cells(f"A2:{chr(ord('A') + const)}2")
        sheet["A2"] = "query"
        sheet[f"{chr(ord('B') + const)}2"] = "frequency_basic"
        sheet[f"{chr(ord('C') + const)}2"] = "freq_accurate"
        sheet[f"{chr(ord('D') + const)}2"] = "heading_entry"
        sheet[f"{chr(ord('E') + const)}2"] = "H1"
        sheet.merge_cells(f"A1:{chr(ord('E') + const)}1")
        # sheet["A1"] = self.url
        header = sheet["A1"]
        header.style = "Note"
        idx = 3
        sheet[f"{chr(ord('A') + cluster_lvl)}{idx + index}"] = cluster
        cluster = self.get_data(cluster)
        sheet[f"{chr(ord('B') + const)}{idx + index}"] = cluster["frequency"]["basic"]
        sheet[f"{chr(ord('C') + const)}{idx + index}"] = cluster["frequency"]["accurate"]
        sheet[f"{chr(ord('D') + const)}{idx + index}"] = cluster["heading_entry"]


    def set_cluster_hard(self, list_in, cluster_lvl):
        frame = pd.DataFrame()
        if len(list_in) == 1 and list_in[0] not in self.blacklist:
            self.cluster_to_excel(list_in[0], cluster_lvl, self.index)
            self.index += 1
            self.blacklist.append(list_in[0])
            return
        for idx, query in enumerate(list_in):
            query_urls = self.get_data(query)["SERP"]["url"]
            tmp_list = []
            clustered = []

            for query1 in list_in[idx + 1:]:
                query_urls1 = self.get_data(query1)["SERP"]["url"]
                if len(set(query_urls) & set(query_urls1)) >= 7 and query1 not in clustered:
                    tmp_list.append(query1)
                    clustered.append(query1)

            if query not in self.blacklist:
                self.cluster_to_excel(query, cluster_lvl, self.index)
                self.index += 1
                self.blacklist.append(query)
            if tmp_list:
                self.set_cluster_hard(clustered, cluster_lvl+1)

    # def set_cluster_hard(self, list_in, cluster_lvl):
    #     count = 0
    #     total = sum(range(0, len(list_in)))
    #     frame = pd.DataFrame()
    #     max_cluster = 0
    #     letter_idx = 0
    #     clustered = []
    #     if len(list_in) == 1:
    #         self.cluster_to_excel(list_in[0], cluster_lvl, self.index)
    #         self.index += 1
    #         print(f'Единственный: {list_in[0]}')
    #         return
    #
    #     for idx, query in enumerate(list_in):
    #         query_urls = self.get_data(query)["SERP"]["url"]
    #         tmp_list = []
    #         flag = False
    #
    #         for query1 in list_in[idx + 1:]:
    #             query_urls1 = self.get_data(query1)["SERP"]["url"]
    #             count += 1        # минимальное количество проверок = сумме всех чисел до значения кол-ва эл-ов
    #             print(f'{count}/{total} Общих URL между {query} и {query1}: '
    #                   f'{len(set(query_urls) & set(query_urls1))}')
    #             if len(set(query_urls) & set(query_urls1)) >= 7 and query1 not in clustered:
    #                 tmp_list.append(query1)
    #                 clustered.append(query1)
    #                 flag = True
    #         if cluster_lvl == 0 and not flag:
    #             list_in.remove(query)
    #         if len(tmp_list) > max_cluster:
    #             for i in range(max_cluster, len(tmp_list)):
    #                 frame = frame.append(pd.Series(), ignore_index=True)
    #             max_cluster = len(tmp_list)
    #         if len(tmp_list) >= 0 and query not in clustered:
    #             frame[f'{query}'] = pd.Series(tmp_list)
    #             letter_idx += 1
    #         if len(tmp_list) == 0:
    #             self.cluster_to_excel(query, cluster_lvl, self.index)
    #             self.index += 1
    #     for cluster in frame:
    #         print(cluster, cluster_lvl)
    #         self.cluster_to_excel(cluster, cluster_lvl, self.index)
    #         self.index += 1
    #         # self.cluster_to_frame(cluster, cluster_lvl)
    #         list_in = self.nan_delete(frame[cluster])  # возвращает лист без NaN
    #         try:
    #             # if frame[cluster][0] == frame[cluster][0]:
    #             self.set_cluster_hard(list_in, cluster_lvl + 1)
    #         except IndexError:
    #             print(f'IndexError при обработке: {cluster}')
    #     return

    # def hard_claster(self, list_):
    #     tmp_list = []
    #     for list_item in list_:
    #         if len(list_item) >= 2:
    #             tmp_list.append(list_item[0])
    #             tmp_list_1 = self.set_cluster(list_item[1:])
    #             tmp = split_list(tmp_list_1)
    #             tmp_list.append(self.hard_claster(tmp))
    #         else:
    #             tmp_list.append(list_item)
    #
    #     return tmp_list

    def get_match(self, list_):
        # tmp = self.hard_claster(tmp)
        out_list = []

        for list_lev2 in list_:
            tmp_d = {}
            if len(list_lev2) > 1:
                tmp_d["query"] = list_lev2[0]
                tmp_d["list_query"] = list_lev2[1:]
            else:
                tmp_d["query"] = list_lev2[0]
                tmp_d["list_query"] = []
            out_list.append(tmp_d)

        return out_list

    # Очищаем список от пустых и не пересекающихся запросов
    # TODO сейчас чистит просто от 1ого элемента
    def clean_query(self, query_list):
        tmp = []
        for item in query_list:
            tmp.append(item)
        data_ = self.generate_out_data(tmp)

        return data_

    # Получение frequency из work_file
    def get_data(self, item):
        for i in self.work_file:
            if item == i["maska"]["with_minsk"]:
                return i

    # Генерация данных в виде кластера
    def generate_out_data(self, cluster):
        tmp = []
        for item in cluster:
            tmp_dict = {"cluster": item["query"],
                        "frequency_basic": int(self.get_data(item["query"])["frequency"]["basic"]),
                        "frequency_accurate": int(self.get_data(item["query"])["frequency"]["accurate"]),
                        "heading_entry": int(self.get_data(item["query"])["heading_entry"]),
                        "queries": []}
            try:
                tmp_dict["H1"] = self.get_data(item["query"])["H1_"]
            except KeyError:
                tmp_dict["H1"] = ""
            for query in item["list_query"]:
                try:
                    in_dict = {"query": query,
                               "frequency_basic": int(self.get_data(query)["frequency"]["basic"]),
                               "frequency_accurate": int(self.get_data(query)["frequency"]["accurate"]),
                               "heading_entry": int(self.get_data(query)["heading_entry"])}
                    try:
                        in_dict["H1"] = self.get_data(query)["H1_"]
                    except KeyError:
                        in_dict["H1"] = " "
                    tmp_dict["queries"].append(in_dict)
                except TypeError:
                    pass
            tmp.append(tmp_dict)
        return tmp

    # сравнивает urls с all_section и при 7+ общих присваивает h1
    def compare_with(self):
        all_section = json_work("other_files/all_section.json", "r")
        for item in self.work_file:
            for item_ in all_section:
                if len(set(item["SERP"]["url"]) & set(item_["SERP"]["url"])) >= 7:
                    item["H1_"] = item_["h1"]

    # создание exel файла
    # def create_excel(self, data_):
    #     num = 0
    #
    #     try:
    #         self.name_doc = os.listdir("excel_files")[-1].split(".")[0]
    #         workbook = load_workbook(filename=f"excel_files/{self.name_doc}.xlsx")
    #     except:
    #         workbook = Workbook(iso_dates=True)
    #
    #     # print(workbook.worksheets)
    #     if len(workbook.worksheets) > 10:
    #         del workbook["Sheet"]
    #         while os.path.exists(f"excel_files/{self.name_doc}.xlsx"):
    #             num += 1
    #             name = self.name_doc.split("_")[0]
    #             self.name_doc = f"{name}_{num}"
    #         else:
    #             workbook = Workbook(iso_dates=True)
    #
    #     sheet = workbook.create_sheet()
    #
    #     const = 2
    #     # sheet["A2"].style = "Good"
    #     # sheet["B2"].style = "Good"
    #     # sheet["C2"].style = "Good"
    #     # sheet["D2"].style = "Good"
    #     # sheet["F2"].style = "Good"
    #     sheet["A2"] = "cluster"
    #     sheet["B2"] = "query"
    #     sheet[f"{chr(ord('C') + const)}2"] = "frequency_basic"
    #     sheet[f"{chr(ord('D') + const)}2"] = "freq_accurate"
    #     sheet[f"{chr(ord('E') + const)}2"] = "heading_entry"
    #     sheet[f"{chr(ord('F') + const)}2"] = "H1"
    #     sheet.merge_cells("A1:F1")
    #     # sheet["A1"] = self.url
    #     header = sheet["A1"]
    #     header.style = "Note"
    #     idx = 3
    #     pred_cluster = None
    #     for i in data_:
    #         idx_ = 0
    #         if i["cluster"] != pred_cluster:
    #             sheet[f"{chr(ord('A') + idx_)}{idx}"] = i["cluster"]
    #             sheet[f"{chr(ord('C') + const)}{idx}"] = i["frequency_basic"]
    #             sheet[f"{chr(ord('D') + const)}{idx}"] = i["frequency_accurate"]
    #             sheet[f"{chr(ord('E') + const)}{idx}"] = i["heading_entry"]
    #             sheet[f"{chr(ord('F') + const)}{idx}"] = i["H1"]
    #             idx += 1
    #             idx_ += 1
    #             try:
    #                 for q in i["queries"]:
    #                     sheet[f"{chr(ord('A') + idx_)}{idx}"] = q["query"]
    #                     sheet[f"{chr(ord('C') + const)}{idx}"] = q["frequency_basic"]
    #                     sheet[f"{chr(ord('D') + const)}{idx}"] = q["frequency_accurate"]
    #                     sheet[f"{chr(ord('E') + const)}{idx}"] = q["heading_entry"]
    #                     sheet[f"{chr(ord('F') + const)}{idx}"] = q["H1"]
    #                     idx += 1
    #                 pred_cluster = i['cluster']
    #             except IndexError:
    #                 pred_cluster = i['cluster']
    #                 idx += 1
    #     workbook.save(filename=f"excel_files/{self.name_doc}.xlsx")
    #     print(f"{sheet.title} добавлен в {self.name_doc}.xlsx")

    # def hard_cluster(self, list_):
    #     tmp_list = []
    #     for query in list_:
    #         if query["queries"]:
    #             tmp_list += query["queries"]
    #             tmp_list += '/'
    #     tmp_list = split_list(tmp_list)
    #     for idx, item in enumerate(tmp_list):
    #         query_list = []
    #         for idx_, query in enumerate(item):
    #             query_list.append(query["query"])

    # запуск скрипта
    def run(self):
        self.compare_with()
        self.get_dict_from_work()  # получаем список запросов в self.list_query
        try:
            self.workbook = create_excel(self.name_doc)
        except FileNotFoundError:
            os.mkdir("excel_files")
            self.workbook = create_excel(self.name_doc)
        self.sheet = self.workbook.create_sheet()
        # tmp_list = self.set_cluster(self.list_query)
        # tmp_list = split_list(tmp_list)
        # clusters = []
        # print("Основные кластеры сформированы!")
        test = ['штукатурка стен минск', 'штукатурка по маякам минск', 'оштукатуривание стен минск',
                'стоимость штукатурки по маякам минск', 'штукатурка стен стоимость минск',
                'оштукатуривание по маякам минск', 'сколько стоит оштукатурить стену минск',
                'сколько стоит штукатурка стен минск', 'на оштукатуривание стен минск']
        # self.set_cluster_hard(test, 0)
        self.set_cluster_hard(self.list_query, 0)

        self.workbook.save(filename=f"excel_files/{self.name_doc}.xlsx")
        print(f"{self.sheet.title} добавлен в {self.name_doc}.xlsx")

        # list_ = self.nan_delete(cluster_frame['декоративная штукатурка минск'])
        # list_ = self.set_cluster_frame(list_, 1)
        # for item in list_:
        #         print(item)

        # for item in tmp_list:
        #     if len(item) > 2:
        #         tmp = self.set_cluster(item[lvl:])
        #         tmp = split_list(tmp)
        #         clusters.append(tmp)
        # clusters = self.get_match(clusters)
        # # print(clusters)
        # query_list = self.get_match(tmp_list)
        # clean_query_list = self.clean_query(query_list)
        # if len(clean_query_list) > 0:
        #     try:
        #         create_excel(self.name_doc)
        #     except FileNotFoundError:
        #         os.mkdir("excel_files")
        #         create_excel(self.name_doc)


if __name__ == "__main__":
    claster = Clasterization(data)
    claster.run()
