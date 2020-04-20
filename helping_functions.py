import json
import os
import re
import string

import requests
from nltk.stem.snowball import SnowballStemmer

from all_constants import API_WORDSTAT, TOKEN_YM, EXCEPT_URLS

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook

''' Вспомогательные функции '''


def tag_to_string(tags):
    strings = []
    for tag in tags:
        strings.append(tag.string)
    return strings


def stemmed(words):
    stemmer = SnowballStemmer("russian")
    list_words = words.split()
    list_out_words = []
    for word in list_words:
        list_out_words.append(stemmer.stem(word))

    return " ".join(list_out_words)


def change_mask(text):
    stopwords = []

    text = text.lower()
    text = re.sub('[%s]' % re.escape(string.punctuation), '', text)  # Удаление пунктуации
    tmp = text.split(" ")
    stem_text = stemmed(text)
    word_list = stem_text.split(" ")
    with open("other_files/stopwords.txt", "r") as f:
        for i in f:
            stopwords.append(i.strip().lower())
    for word in stopwords:
        word = stemmed(word)
        for word_ in word_list:
            if word == word_:
                # print("до обработки:", tmp)
                # print(word, "==", word_)
                try:
                    tmp.pop(word_list.index(word_))
                except IndexError:
                    tmp.pop(word_list.index(word_) - 1)
                # print("после обработки:", tmp)
    return " ".join(tmp)


def check_except_url(url):
    status = True
    for i in EXCEPT_URLS:
        if i == url:
            status = False

    return status


def check_geo(text):
    list_ = []
    status = True
    with open("other_files/geo.txt", "r") as f:
        for i in f:
            list_.append(i.strip().lower())
    for word in list_:
        if word in text:
            status = False

    return status


def masked(text):
    maska = {}
    text = text.replace('-', ' ')  # заменяем '-' на пробел
    text = re.sub(r"\((.*?)\)", "", text)   # удаляем скобки и их содержание
    text = change_mask(text).strip()
    with_minsk = text.replace('в минске', 'минск')
    if with_minsk.count("минск") < 1 and check_geo(with_minsk):
        with_minsk += ' минск'
    elif with_minsk.count("минск") > 1:
        with_minsk = with_minsk.replace('минск', '')
    without_minsk = with_minsk.replace('минск', '')
    maska["with_minsk"] = with_minsk
    maska["without_minsk"] = without_minsk
    return maska


def split_list(list_in, delimiter="/"):
    list_ = []
    tmp = []
    for i in list_in:
        if i != delimiter:
            list_.append(i)
        else:
            if len(list_) > 0:
                tmp.append(list_)
            list_ = []
    return tmp


def json_work(filename, method, write_data=None):
    if write_data is None:
        write_data = []

    with open(filename, method, encoding="utf-8") as f:
        if method == "r":
            data = json.load(f)
            return data
        elif method == "w":
            json.dump(write_data, f, ensure_ascii=False, indent=2)


def get_request_to_ya(data):
    jdata = json.dumps(data, ensure_ascii=False).encode('utf-8')
    try:
        response = requests.post(API_WORDSTAT, jdata, headers={"Authorization": f"Bearer {TOKEN_YM}"})
        resp_json = response.json()
        return resp_json
    except requests.exceptions.ConnectionError:
        return get_request_to_ya(data)


def create_excel(name_doc):
    num = 0

    try:
        workbook = load_workbook(filename=f"excel_files/{name_doc}.xlsx")
    except:
        workbook = Workbook(iso_dates=True)
    if len(workbook.worksheets) > 10:
        for sheetName in workbook.sheetnames:
            del workbook[sheetName]
    return workbook

def set_filename(name_doc):
    num = 0
    try:
        name_doc = os.listdir("excel_files")[-1].split(".")[0]
        workbook = load_workbook(filename=f"excel_files/{name_doc}.xlsx")
        if len(workbook.worksheets) > 10:
            del workbook["Sheet"]
            while os.path.exists(f"excel_files/{name_doc}.xlsx"):
                num += 1
                name = name_doc.split("_")[0]
                name_doc = f"{name}_{num}"
        return name_doc
    except:
        return name_doc

