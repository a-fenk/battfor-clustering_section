import os
from pprint import pprint

from openpyxl.styles import PatternFill, Fill


from helping_functions import json_work, create_excel, set_filename, check_description

# data = json_work("other_files/main.json", "r")


class Clustering:
    def __init__(self, work_file, url, name_doc="file_"):
        self.index = 0
        self.workbook = None
        self.sheet = None
        self.work_file = work_file
        self.list_query = []
        self.name_doc = name_doc
        self.blacklist = []
        self.url = url

    # Получаем список запросов
    def get_dict_from_work(self):
        for item in self.work_file:
            self.list_query.append(item["maska"]["with_minsk"])

    def cluster_to_excel(self, item, cluster_lvl, index, color):
        const = 3
        sheet = self.sheet
        sheet.merge_cells(f"A1:{chr(ord('F') + const)}1")
        sheet["A1"] = self.url
        sheet["A2"] = "H1"
        sheet["B2"] = "Новые ключи"
        sheet.merge_cells(f"C2:{chr(ord('C') + const)}2")
        sheet["C2"] = "query"
        sheet[f"{chr(ord('D') + const)}2"] = "frequency_basic"
        sheet[f"{chr(ord('E') + const)}2"] = "freq_accurate"
        sheet[f"{chr(ord('F') + const)}2"] = "heading_entry"
        header = sheet["A1"]
        header.style = "Note"
        idx = 3
        if color:
            sheet[f"{chr(ord('C') + cluster_lvl)}{idx + index}"].fill = PatternFill("solid", fgColor="9f9f9f")
        sheet.merge_cells(f"{chr(ord('C') + cluster_lvl)}{idx + index}:F{idx + index}")
        sheet[f"{chr(ord('C') + cluster_lvl)}{idx + index}"] = item
        item = self.get_data(item)
        if cluster_lvl == 0:
            fgcolor = check_description(item["SERP"]["description"])
            sheet[f"{chr(ord('C') + cluster_lvl)}{idx + index}"].fill = PatternFill("solid", fgColor=fgcolor)
        sheet[f"{chr(ord('D') + const)}{idx + index}"] = item["frequency"]["basic"]
        sheet[f"{chr(ord('E') + const)}{idx + index}"] = item["frequency"]["accurate"]
        sheet[f"{chr(ord('F') + const)}{idx + index}"] = item["heading_entry"]
        sheet[f"A{idx + index}"] = item["H1"]

    def set_cluster_hard(self, list_in, cluster_lvl):
        if len(list_in) == 1 and list_in[0] not in self.blacklist:
            self.cluster_to_excel(list_in[0], cluster_lvl, self.index, False)
            self.index += 1
            self.blacklist.append(list_in[0])
            return
        index = -1
        for idx, query in enumerate(list_in):
            if query not in self.blacklist:
                query_urls = self.get_data(query)["SERP"]["url"]
                tmp_list = []
                clustered = []
                flag = False
                if cluster_lvl == 0:
                    index += 1
                else:
                    index = idx
                for query1 in list_in[idx + 1:]:
                    query_urls1 = self.get_data(query1)["SERP"]["url"]
                    if query1 not in self.blacklist:
                        if len(set(query_urls) & set(query_urls1)) == 10:
                            if query not in self.blacklist:
                                self.cluster_to_excel(query, cluster_lvl, self.index, False)
                                self.index += 1
                                self.blacklist.append(query)
                                flag = True
                            if cluster_lvl < 3 and flag:
                                cluster_lvl_ = cluster_lvl+1
                            else:
                                cluster_lvl_ = cluster_lvl
                            self.cluster_to_excel(query1, cluster_lvl_, self.index, True)
                            self.index += 1
                            self.blacklist.append(query1)
                        elif cluster_lvl == 3:
                            self.cluster_to_excel(query1, cluster_lvl, self.index, False)
                            self.index += 1
                            self.blacklist.append(query1)
                    if len(set(query_urls) & set(query_urls1)) >= 7 and query1 not \
                            in (clustered or self.blacklist) and cluster_lvl < 3:
                        tmp_list.append(query1)
                        clustered.append(query1)

                if query not in self.blacklist:
                    self.cluster_to_excel(query, cluster_lvl, self.index, False)
                    self.index += 1
                    self.blacklist.append(query)
                if tmp_list:
                    self.set_cluster_hard(clustered, cluster_lvl + 1)
        return

    # Получение frequency из work_file
    def get_data(self, item):
        for i in self.work_file:
            if item == i["maska"]["with_minsk"]:
                return i

    # сравнивает urls с all_section и при 7+ общих присваивает h1
    def compare_with(self):
        all_section = json_work("other_files/all_section.json", "r")
        tmp = []
        for item in self.work_file:
            idx = 0
            flag = False
            while idx < len(all_section) and not flag:
                if len(set(item["SERP"]["url"]) & set(all_section[idx]["SERP"]["url"])) >= 7:
                    item["H1"] = all_section[idx]["h1"]
                    flag = True
                else:
                    item["H1"] = ""
                idx += 1
            tmp.append(item)
        json_work("other_files/work_file.json", "w", tmp)

    # запуск скрипта
    def run(self):
        self.compare_with()
        self.get_dict_from_work()  # получаем список запросов в self.list_query
        try:
            self.name_doc = set_filename(self.name_doc)
            self.workbook = create_excel(self.name_doc)
            self.sheet = self.workbook.create_sheet()
            self.set_cluster_hard(self.list_query, 0)
            self.workbook.save(filename=f"excel_files/{self.name_doc}.xlsx")
        except FileNotFoundError:
            os.mkdir("excel_files")
            self.workbook = create_excel(self.name_doc)
            self.sheet = self.workbook.create_sheet()
            self.set_cluster_hard(self.list_query, 0)
            self.workbook.save(filename=f"excel_files/{self.name_doc}.xlsx")

        print(f"{self.sheet.title} добавлен в {self.name_doc}.xlsx")


# if __name__ == "__main__":
#     cluster = Clustering(data)
#     cluster.run()
